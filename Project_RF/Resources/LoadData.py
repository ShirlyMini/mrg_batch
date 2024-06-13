# [[u1, p1, e1],[u2,p2,e2]...]
import openpyxl
import json
import yaml

def fetch_data_xl(xl_path):
    wb=openpyxl.load_workbook(xl_path)
    sh = wb['Sheet1']
    print(sh.max_row)#5
    print(sh.max_column)#3
    l2=[]
    for r in range(2, sh.max_row+1): # 0 to 4
        l1=[]
        for c in range(1, sh.max_column+1):
            #print(sh.cell(r,c).value)
            l1.append(sh.cell(r,c).value)
        #print(l1)#['admin@yourstore.com', 'admin', 'Pass']
        l2.append(l1)
    print(l2)
    return l2

def fetch_data_json(json_path):
    fr=open(json_path)
    json_content=fr.read()

    json_dict=json.loads(json_content)
    # print(json_dict)
    # print(json_dict["login_data"])
    # print(json_dict["login_data"]["row1"])
    # print(list(json_dict["login_data"].values()))
    return list(json_dict["login_data"].values())

def fetch_data_yaml(yaml_path):
    fr=open(yaml_path)
    yaml_content=fr.read()

    yaml_dict = yaml.load(yaml_content, Loader=yaml.FullLoader)
    print(yaml_dict)
    print(list(yaml_dict["login_data"].values()))
    return list(yaml_dict["login_data"].values())



# fetch_data_json("../Data/logindata_json.json")
# fetch_data_yaml("../Data/test_data_yaml.yml")
# func call
# fetch_data_xl("../Data/LoginData.xlsx")
