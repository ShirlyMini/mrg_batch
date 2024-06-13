import openpyxl

## open xl/wb-sheet-row and col-cell-fetch the data

wb=openpyxl.load_workbook(r".\LoginData.xlsx")
sheet=wb['Sheet1']
# print(sheet.cell(1,1).value)
# print(sheet.cell(1,2).value)
# print(sheet.cell(1,3).value)
# print(sheet.max_row)# 5
# print(sheet.max_column) # 3

# for i in range(sheet.max_row):# 0 to 4
for row in range(2,sheet.max_row+1):# 1 to 5
    print(sheet.cell(row,1).value, sheet.cell(row,2).value,sheet.cell(row,3).value)